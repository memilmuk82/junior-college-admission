from __future__ import annotations

import csv
import hashlib
import re
from collections.abc import Iterable, Mapping
from pathlib import Path

from openpyxl import load_workbook

SOURCE_SIZE = 394_539
SOURCE_SHA256 = "bde8fe5d513ce2737c08815b0d7e1df366dc8844e6ff7f243eccb63c3bd40606"
SOURCE_PATH = Path(
    "tmp/codex-reference/xlsx/OOO_2026 전문대 수시 분석_2027 수도권 전문대 입결_전문대교협.xlsx"
)
SUPPLEMENTAL_SIZE = 5_671_593
SUPPLEMENTAL_SHA256 = "c35d548abb244168dcffd1a582a38368a58f611f2998604a9b1b70f7e5ae6658"
SUPPLEMENTAL_PATH = Path("tmp/codex-reference/csv/procollege_all_2026 (1).csv")
CATALOG_PATH = Path("data/seed/phase14_public_admission_catalog.csv")
RESULT_PATH = Path("data/seed/phase14_public_admission_results_2025.csv")

INSTITUTION_CODES = {
    "동양미래대학교": "DONGYANG-MIRAE",
    "명지전문대학": "MYONGJI-COLLEGE",
    "인하공업전문대학": "INHA-TECHNICAL-COLLEGE",
    "연성대학교": "YEONSUNG",
}
ROUND_CODES = {"수시1차": "SUSI-1", "수시2차": "SUSI-2"}
TRACK_CODES = {
    ("특별전형", "일반고"): "SPECIAL-GENERAL-HS",
    ("특별전형", "특성화고"): "SPECIAL-VOCATIONAL-HS",
    ("일반전형", "일반전형"): "GENERAL",
    ("특별전형", "대학자체"): "COLLEGE-SPECIFIC",
}
PUBLIC_RESULT_SHEETS = ("2025 수시(1차) 결과", "2025 수시(2차) 결과")


def main() -> None:
    source = SOURCE_PATH.read_bytes()
    if len(source) != SOURCE_SIZE or hashlib.sha256(source).hexdigest() != SOURCE_SHA256:
        raise SystemExit("BLOCKED_SOURCE: 기준 XLSX size 또는 SHA-256이 다릅니다.")
    supplemental_source = SUPPLEMENTAL_PATH.read_bytes()
    if (
        len(supplemental_source) != SUPPLEMENTAL_SIZE
        or hashlib.sha256(supplemental_source).hexdigest() != SUPPLEMENTAL_SHA256
    ):
        raise SystemExit("BLOCKED_SOURCE: 보조 CSV size 또는 SHA-256이 다릅니다.")

    supplemental = _supplemental_index()

    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    catalog: dict[tuple[str, ...], dict[str, object]] = {}
    results: list[dict[str, object]] = []
    try:
        for sheet_name in PUBLIC_RESULT_SHEETS:
            worksheet = workbook[sheet_name]
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=3, max_col=11, values_only=True), start=3
            ):
                source_institution = _text(row[2])
                institution_name = re.sub(r"\s*\(특\)\s*$", "", source_institution)
                institution_code = INSTITUTION_CODES.get(institution_name)
                if institution_code is None:
                    continue
                region = _text(row[1])
                round_name = _text(row[3])
                program_name = _text(row[4])
                capacity = _number_or_blank(row[5])
                day_night = _text(row[6])
                category = _text(row[7])
                track_label = _text(row[8])
                cutoff = _number_or_blank(row[9])
                average = _number_or_blank(row[10])
                track_code = TRACK_CODES[(category, track_label)]
                program_code = _program_code(program_name)
                round_code = ROUND_CODES[round_name]
                track_name = f"{category} / {track_label}"
                catalog_key = (institution_code, program_code, round_code, track_code)
                if catalog_key in catalog:
                    raise SystemExit(
                        f"canonical 업무키가 중복되었습니다: {sheet_name}:{row_number}"
                    )
                catalog[catalog_key] = {
                    "institution_code": institution_code,
                    "institution_name": institution_name,
                    "institution_type": "JUNIOR_COLLEGE",
                    "campus_code": "MAIN",
                    "campus_name": "본교",
                    "program_code": program_code,
                    "program_name": program_name,
                    "target_academic_year": 2027,
                    "admission_round_code": round_code,
                    "admission_round_name": round_name,
                    "admission_track_code": track_code,
                    "admission_track_name": track_name,
                    "mapping_status": "EXPLICIT_PHASE14_PILOT",
                }
                supplemental_match = supplemental.get(catalog_key)
                competition_rate = (
                    supplemental_match[1].get("경쟁률", "") if supplemental_match else ""
                )
                source_reference = f"sha256:{SOURCE_SHA256}#{sheet_name}:{row_number}"
                if competition_rate:
                    assert supplemental_match is not None
                    source_reference += (
                        ";supplemental.competition_rate="
                        f"sha256:{SUPPLEMENTAL_SHA256}#CSV:{supplemental_match[0]}"
                    )
                results.append(
                    {
                        "결과학년도": 2025,
                        "지역": region,
                        "대학명": institution_name,
                        "캠퍼스명": "본교",
                        "모집시기": round_name,
                        "전공명": program_name,
                        "주야": day_night,
                        "전형구분1": category,
                        "전형구분2": track_label,
                        "모집인원": capacity,
                        "지원자수": "",
                        "합격자수": "",
                        "경쟁률": competition_rate,
                        "합격자최고": "",
                        "합격자평균": average,
                        "합격자최저": cutoff,
                        "점수기준": "RANK_GRADE",
                        "점수방향": "LOWER_IS_BETTER",
                        "source_reference": source_reference,
                    }
                )
    finally:
        workbook.close()

    if len(results) != 482 or len(catalog) != 482:
        raise SystemExit(
            f"공개 파일럿 결과 행/업무키 수가 예상과 다릅니다: {len(results)}/{len(catalog)}"
        )
    if sum(bool(row["경쟁률"]) for row in results) != 482:
        raise SystemExit("보조 CSV exact 업무키 경쟁률 보강 수가 482행과 일치하지 않습니다.")
    _write_csv(CATALOG_PATH, tuple(catalog.values()))
    _write_csv(RESULT_PATH, results)
    print(
        f"public seed rows={len(results)} catalog_keys={len(catalog)} source_sha256={SOURCE_SHA256}"
    )


def _write_csv(path: Path, rows: Iterable[Mapping[str, object]]) -> None:
    materialized = list(rows)
    if not materialized:
        raise SystemExit(f"빈 seed를 만들 수 없습니다: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(materialized[0]), lineterminator="\n")
        writer.writeheader()
        writer.writerows(materialized)


def _text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _number_or_blank(value: object) -> object:
    return value if value is not None else ""


def _supplemental_index() -> dict[tuple[str, ...], tuple[int, dict[str, str]]]:
    grouped: dict[tuple[str, ...], list[tuple[int, dict[str, str]]]] = {}
    with SUPPLEMENTAL_PATH.open(encoding="utf-8-sig", newline="") as handle:
        for row_number, row in enumerate(csv.DictReader(handle), start=2):
            institution_name = _text(row.get("대학명"))
            round_name = _text(row.get("모집시기"))
            category = _text(row.get("전형구분"))
            track_label = _text(row.get("출신교"))
            institution_code = INSTITUTION_CODES.get(institution_name)
            if (
                _text(row.get("모집학년도")) != "2025"
                or institution_code is None
                or round_name not in ROUND_CODES
                or (category, track_label) not in TRACK_CODES
            ):
                continue
            key = (
                institution_code,
                _program_code(_text(row.get("전공명"))),
                ROUND_CODES[round_name],
                TRACK_CODES[(category, track_label)],
            )
            grouped.setdefault(key, []).append((row_number, row))
    return {key: rows[0] for key, rows in grouped.items() if len(rows) == 1}


def _program_code(name: str) -> str:
    digest = hashlib.sha256(name.encode("utf-8")).hexdigest()[:12].upper()
    return f"PROGRAM-{digest}"


if __name__ == "__main__":
    main()
