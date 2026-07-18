from __future__ import annotations

import argparse
import hashlib
from pathlib import Path

from openpyxl import load_workbook

EXPECTED_SIZE = 394_539
EXPECTED_SHA256 = "bde8fe5d513ce2737c08815b0d7e1df366dc8844e6ff7f243eccb63c3bd40606"
EXPECTED_SHEETS = (
    "성적 입력",
    "대학별 등급",
    "2025 수시(1차) 결과",
    "2025 수시(2차) 결과",
)
EXPECTED_PUBLIC_RESULT_ROWS = {
    "2025 수시(1차) 결과": 1_818,
    "2025 수시(2차) 결과": 1_652,
}
SUPPORTED_FORMULA_CELLS = (
    "J8",
    "J9",
    "J10",
    "J11",
    "J15",
    "J16",
    "J17",
    "J19",
    "J21",
    "J23",
    "J24",
)
KNOWN_BLOCKED_FORMULAS = {
    "J18": ("$C$4:$G$4", "$E$8:$G$8"),
    "J20": ("$E$4:$F$4", "$F$4*0.2"),
}
RESULT_HEADERS = (
    "지역",
    "대학명",
    "모집시기",
    "전공명",
    "모집시기별입학정원",
    "주/야",
    "전형구분1",
    "전형구분2",
    "합격자최저",
    "합격자평균",
)


class ReferenceVerificationError(RuntimeError):
    pass


def _digest(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalized_header(value: object) -> str:
    return "" if value is None else "".join(str(value).split())


def verify_reference(path: Path) -> dict[str, object]:
    if not path.is_file():
        raise ReferenceVerificationError("기준 XLSX를 찾을 수 없습니다.")
    size = path.stat().st_size
    digest = _digest(path)
    if size != EXPECTED_SIZE or digest != EXPECTED_SHA256:
        raise ReferenceVerificationError("BLOCKED_SOURCE: 기준 XLSX size 또는 SHA-256 불일치")

    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    try:
        if tuple(workbook.sheetnames) != EXPECTED_SHEETS:
            raise ReferenceVerificationError("BLOCKED_SOURCE: 기준 XLSX 시트 구성이 다릅니다.")

        published_counts: dict[str, int] = {}
        for sheet_name, expected_count in EXPECTED_PUBLIC_RESULT_ROWS.items():
            sheet = workbook[sheet_name]
            headers = tuple(
                _normalized_header(sheet.cell(2, column).value) for column in range(2, 12)
            )
            if headers != RESULT_HEADERS:
                raise ReferenceVerificationError(f"BLOCKED_SOURCE: {sheet_name} 머리글 불일치")
            count = sum(
                1
                for row in range(3, sheet.max_row + 1)
                if any(sheet.cell(row, column).value is not None for column in range(2, 12))
            )
            if count != expected_count:
                raise ReferenceVerificationError(
                    f"BLOCKED_SOURCE: {sheet_name} 공개 결과 행 수 {count}, 기대 {expected_count}"
                )
            published_counts[sheet_name] = count

        formulas = workbook["대학별 등급"]
        for cell in SUPPORTED_FORMULA_CELLS:
            formula = formulas[cell].value
            if not isinstance(formula, str) or not formula.startswith("="):
                raise ReferenceVerificationError(f"BLOCKED_SOURCE: 지원 수식 {cell} 누락")
        for cell, required_fragments in KNOWN_BLOCKED_FORMULAS.items():
            formula = formulas[cell].value
            if not isinstance(formula, str) or any(
                fragment not in formula for fragment in required_fragments
            ):
                raise ReferenceVerificationError(f"BLOCKED_SOURCE: 알려진 차단 수식 {cell} 변경")

        return {
            "size": size,
            "sha256": digest,
            "sheets": tuple(workbook.sheetnames),
            "public_result_rows": published_counts,
            "supported_formula_cells": SUPPORTED_FORMULA_CELLS,
            "blocked_formula_cells": tuple(KNOWN_BLOCKED_FORMULAS),
            "student_input_values_emitted": False,
        }
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Phase 14 기준 XLSX 비식별 계약 검증")
    parser.add_argument("path", type=Path)
    args = parser.parse_args()
    try:
        result = verify_reference(args.path)
    except ReferenceVerificationError as error:
        print(str(error))
        return 2
    counts = result["public_result_rows"]
    assert isinstance(counts, dict)
    sheets = result["sheets"]
    assert isinstance(sheets, tuple)
    print(
        "PASS_REFERENCE_XLSX "
        f"size={result['size']} sha256={result['sha256']} "
        f"sheets={len(sheets)} "
        f"first={counts['2025 수시(1차) 결과']} "
        f"second={counts['2025 수시(2차) 결과']} "
        "blocked=J18,J20 student_input_values_emitted=false"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
