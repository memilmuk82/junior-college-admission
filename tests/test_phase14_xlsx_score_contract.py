from __future__ import annotations

from dataclasses import replace
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services.score_calculation import calculate_reflected_grade
from app.services.score_inputs import (
    AcademicRecordInput,
    CourseRecordInput,
    GradeSourcePolicy,
    ScoreInputSelection,
    ScoreInputStatus,
    ScoreInputTrace,
)
from app.services.score_selection import ComparableCourseValue, select_terms_and_subjects
from scripts.verify_phase14_reference_xlsx import verify_reference
from tests.test_score_selection import _definition

SEMESTER_GRADES = {
    (1, 1): Decimal("2"),
    (1, 2): Decimal("5"),
    (2, 1): Decimal("3"),
    (2, 2): Decimal("4"),
    (3, 1): Decimal("1"),
}


def _source(
    terms: tuple[tuple[int, int], ...],
    semester_grades: dict[tuple[int, int], Decimal] = SEMESTER_GRADES,
) -> ScoreInputSelection:
    records = tuple(
        AcademicRecordInput(
            academic_record_id=f"xlsx-record-{grade}-{semester}",
            academic_year=2024 + grade,
            grade=grade,
            semester=semester,
            record_source="HOME_SCHOOL_RECORD",
            is_vocational_training_semester=False,
            verification_status="USER_VERIFIED",
            courses=(
                CourseRecordInput(
                    course_record_id=f"xlsx-course-{grade}-{semester}",
                    subject_group="합성 교과",
                    subject_name=f"합성 {grade}-{semester}",
                    credits=Decimal("1"),
                    raw_score=None,
                    raw_score_label=None,
                    course_mean=None,
                    standard_deviation=None,
                    achievement_level=None,
                    enrollment_count=None,
                    rank_grade=semester_grades[(grade, semester)],
                    user_verified=True,
                ),
            ),
        )
        for grade, semester in terms
    )
    return ScoreInputSelection(
        status=ScoreInputStatus.READY,
        records=records,
        trace=ScoreInputTrace(
            rule_id="xlsx-contract-scope",
            rule_version="phase14-v1",
            policy=GradeSourcePolicy.HOME_ONLY,
            selected_sources=("HOME_SCHOOL_RECORD",),
            selected_terms=(),
            exclusion_reasons=(),
        ),
    )


def _calculate(cell: str, semester_grades: dict[tuple[int, int], Decimal] = SEMESTER_GRADES):  # type: ignore[no-untyped-def]
    all_terms = tuple(semester_grades)
    first_four = all_terms[:4]
    definitions = {
        "J8": (
            first_four,
            replace(_definition(), semester_selection_method="BEST_N", best_semester_count=2),
        ),
        "J9": (
            all_terms,
            replace(_definition(), semester_selection_method="BEST_N", best_semester_count=2),
        ),
        "J10": (
            first_four,
            replace(_definition(), semester_selection_method="BEST_N", best_semester_count=1),
        ),
        "J11": (
            all_terms,
            replace(_definition(), semester_selection_method="BEST_N", best_semester_count=1),
        ),
        "J15": (
            all_terms,
            replace(
                _definition(),
                semester_selection_method="BEST_N",
                semester_selection_scope="PER_GRADE",
                best_semester_count=1,
                weighting_mode="GRADE_ONLY",
                grade_weight_1=Decimal("0.30"),
                grade_weight_2=Decimal("0.30"),
                grade_weight_3=Decimal("0.40"),
            ),
        ),
        "J16": (
            all_terms,
            replace(
                _definition(),
                weighting_mode="GRADE_ONLY",
                grade_weight_1=Decimal("0.30"),
                grade_weight_2=Decimal("0.30"),
                grade_weight_3=Decimal("0.40"),
            ),
        ),
        "J17": (
            first_four,
            replace(
                _definition(),
                weighting_mode="GRADE_ONLY",
                grade_weight_1=Decimal("0.40"),
                grade_weight_2=Decimal("0.60"),
            ),
        ),
        "J19": (
            all_terms,
            replace(
                _definition(),
                weighting_mode="GRADE_ONLY",
                grade_weight_1=Decimal("0.20"),
                grade_weight_2=Decimal("0.30"),
                grade_weight_3=Decimal("0.50"),
            ),
        ),
        "J21": (
            first_four,
            replace(
                _definition(),
                weighting_mode="GRADE_ONLY",
                grade_weight_1=Decimal("0.50"),
                grade_weight_2=Decimal("0.50"),
            ),
        ),
        "J23": (
            first_four,
            replace(_definition(), semester_selection_method="BEST_N", best_semester_count=3),
        ),
        "J24": (all_terms[:2], _definition()),
    }
    terms, definition = definitions[cell]
    definition = replace(definition, value_direction="LOWER_IS_BETTER")
    values = {
        f"xlsx-course-{grade}-{semester}": ComparableCourseValue(
            course_record_id=f"xlsx-course-{grade}-{semester}",
            normalized_value=semester_grades[(grade, semester)],
            value_scale="RANK_GRADE",
            scope_codes=frozenset({"GENERAL_SUBJECTS"}),
        )
        for grade, semester in terms
    }
    selected = select_terms_and_subjects(_source(terms, semester_grades), definition, values)
    return calculate_reflected_grade(
        selected,
        definition,
        rule_id=f"xlsx-contract-{cell}",
        rule_version="phase14-v1",
    )


@pytest.mark.parametrize(
    ("cell", "expected"),
    (
        ("J8", "2.50"),
        ("J9", "1.50"),
        ("J10", "2.00"),
        ("J11", "1.00"),
        ("J15", "1.90"),
        ("J16", "2.50"),
        ("J17", "3.50"),
        ("J19", "2.25"),
        ("J21", "3.50"),
        ("J23", "3.00"),
        ("J24", "3.50"),
    ),
)
def test_supported_xlsx_semester_formulas_match_limited_score_engine(
    cell: str, expected: str
) -> None:
    result = _calculate(cell)

    assert result.display_average_grade == Decimal(expected)
    assert result.trace.rule_id == f"xlsx-contract-{cell}"
    assert result.trace.rounding_stage == "DISPLAY_ONLY"


def test_xlsx_grade_weighted_contract_exposes_terms_weights_and_rounding_trace() -> None:
    result = _calculate("J15")

    assert [(row.grade, row.semester) for row in result.trace.selected_semesters] == [
        (1, 1),
        (2, 1),
        (3, 1),
    ]
    assert [(row.key, row.value, row.weight) for row in result.trace.components] == [
        ("grade_1", Decimal("2"), Decimal("0.30")),
        ("grade_2", Decimal("3"), Decimal("0.30")),
        ("grade_3", Decimal("1"), Decimal("0.40")),
    ]
    assert result.unrounded_average_grade == Decimal("1.90")
    assert result.trace.rounding_mode == "ROUND_HALF_UP"
    assert result.trace.rounding_scale == 2


def test_supported_cells_match_reference_xlsx_cached_values_without_emitting_student_values() -> (
    None
):
    path = Path(
        "tmp/codex-reference/xlsx/OOO_2026 전문대 수시 분석_2027 수도권 전문대 입결_전문대교협.xlsx"
    )
    if not path.is_file():
        pytest.skip("읽기 전용 기준 XLSX가 현재 환경에 없습니다.")
    verify_reference(path)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet = workbook["대학별 등급"]
        terms = tuple(SEMESTER_GRADES)
        semester_grades: dict[tuple[int, int], Decimal] = {}
        for term, column in zip(terms, range(3, 8), strict=True):
            value = sheet.cell(4, column).value
            if not isinstance(value, (int, float, Decimal)):
                pytest.fail(f"기준 XLSX {sheet.cell(4, column).coordinate} 캐시 누락")
            semester_grades[term] = Decimal(str(value))
        for cell in ("J8", "J9", "J10", "J11", "J15", "J16", "J17", "J19", "J21", "J23", "J24"):
            cached = sheet[cell].value
            if not isinstance(cached, (int, float, Decimal)):
                pytest.fail(f"기준 XLSX {cell} 캐시 누락")
            expected = Decimal(str(cached)).quantize(Decimal("0.01"))
            actual = _calculate(cell, semester_grades).display_average_grade
            if actual != expected:
                pytest.fail(f"기준 XLSX {cell} 캐시와 앱 결과 불일치")
    finally:
        workbook.close()
