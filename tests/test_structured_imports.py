from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from app.services.structured_imports import (
    parse_structured_text,
    parse_target_student_text,
    parse_target_student_xlsx_bytes,
    parse_xlsx_bytes,
)


def _synthetic_workbook_bytes() -> bytes:
    workbook = Workbook()
    first_sheet = workbook.active
    assert first_sheet is not None
    first_sheet.title = "합성 원적교"
    first_sheet.append(["합성 입력 설명"])
    first_sheet.append([])
    first_sheet.append(["학년도", "학년", "학기", "과목", "원점수"])
    first_sheet.append([2026, 2, 1, "합성 과목 A", 91])

    second_sheet = workbook.create_sheet("합성 위탁기관")
    second_sheet.append(["학년", "학기", "과목", "성취도"])
    second_sheet.append([3, 1, "합성 과목 B", "P"])

    ignored_sheet = workbook.create_sheet("머리글 없음")
    ignored_sheet.append(["합성 안내만 있음"])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_csv_normalizes_known_fields_and_preserves_pass_value() -> None:
    source = "\n".join(
        [
            "학년도,학년,학기,교과,과목,이수단위,원점수,평균,표준편차,성취도,수강자수,석차등급",
            "2026,2,1,수학,합성 과목,3,P,,,,,",
        ]
    )

    preview = parse_structured_text(source, source_format="csv")

    assert preview.rows[0].academic_year == 2026
    assert preview.rows[0].credits == Decimal("3")
    assert preview.rows[0].raw_score == "P"
    assert preview.rows[0].course_mean is None
    assert preview.rows[0].rank_grade is None
    assert preview.issues == ()


def test_pasted_table_detects_tabs_and_reports_invalid_number_without_zeroing() -> None:
    source = "학년\t학기\t과목\t원점수\n3\t2\t합성 과목\t확인필요"

    preview = parse_structured_text(source, source_format="pasted_table")

    assert preview.rows[0].raw_score is None
    assert preview.rows[0].subject_name == "합성 과목"
    assert [(issue.code, issue.field, issue.row_number) for issue in preview.issues] == [
        ("INVALID_DECIMAL", "raw_score", 2)
    ]


def test_same_source_produces_same_sha256_without_retaining_unknown_columns() -> None:
    source = "학년,학기,과목,검수제외열\n1,1,합성 과목,저장하지 않음"

    first = parse_structured_text(source, source_format="csv")
    second = parse_structured_text(source, source_format="csv")

    assert first.source_hash == second.source_hash
    assert len(first.source_hash) == 64
    assert not hasattr(first.rows[0], "검수제외열")
    assert first.ignored_headers == ("검수제외열",)


def test_extra_cells_are_not_retained_or_allowed_to_break_preview() -> None:
    source = "학년,학기,과목\n1,1,합성 과목,저장하지 않음"

    preview = parse_structured_text(source, source_format="csv")

    assert len(preview.rows) == 1
    assert preview.rows[0].subject_name == "합성 과목"


def test_xlsx_detects_headers_after_preamble_and_reads_multiple_sheets() -> None:
    source = _synthetic_workbook_bytes()

    preview = parse_xlsx_bytes(source)

    assert [
        (row.source_sheet, row.source_row_number, row.subject_name) for row in preview.rows
    ] == [
        ("합성 원적교", 4, "합성 과목 A"),
        ("합성 위탁기관", 2, "합성 과목 B"),
    ]
    assert preview.rows[0].raw_score == Decimal("91")
    assert preview.rows[1].achievement_level == "P"
    assert preview.source_hash == parse_xlsx_bytes(source).source_hash


def test_xlsx_reports_sheet_without_recognizable_header() -> None:
    preview = parse_xlsx_bytes(_synthetic_workbook_bytes())

    assert [(issue.code, issue.sheet_name, issue.row_number) for issue in preview.issues] == [
        ("HEADER_NOT_FOUND", "머리글 없음", 0)
    ]


def test_class_table_returns_only_target_student_rows_and_course_statistics() -> None:
    source = "\n".join(
        [
            "학생식별자,학년,학기,과목,원점수,평균,표준편차,수강자수",
            "SYNTHETIC_OTHER,3,1,합성 과목,77,82.5,4.2,20",
            "SYNTHETIC_TARGET,3,1,합성 과목,91,82.5,4.2,20",
        ]
    )

    preview = parse_target_student_text(
        source,
        source_format="csv",
        target_student_reference="SYNTHETIC_TARGET",
    )

    assert len(preview.rows) == 1
    assert preview.rows[0].raw_score == Decimal("91")
    assert preview.rows[0].course_mean == Decimal("82.5")
    assert preview.rows[0].standard_deviation == Decimal("4.2")
    assert preview.rows[0].enrollment_count == 20
    assert "SYNTHETIC_OTHER" not in repr(preview)
    assert "SYNTHETIC_TARGET" not in repr(preview)


def test_class_table_reports_missing_target_without_returning_other_rows() -> None:
    source = "학생식별자,과목,원점수\nSYNTHETIC_OTHER,합성 과목,77"

    preview = parse_target_student_text(
        source,
        source_format="csv",
        target_student_reference="SYNTHETIC_TARGET",
    )

    assert preview.rows == ()
    assert [(issue.code, issue.field) for issue in preview.issues] == [
        ("TARGET_NOT_FOUND", "student_reference")
    ]
    assert "SYNTHETIC_OTHER" not in repr(preview)


def test_xlsx_class_table_returns_only_target_student_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "합성 학급"
    worksheet.append(["학생식별자", "과목", "원점수", "평균", "수강자수"])
    worksheet.append(["SYNTHETIC_OTHER", "합성 과목", 77, 82.5, 20])
    worksheet.append(["SYNTHETIC_TARGET", "합성 과목", 91, 82.5, 20])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    preview = parse_target_student_xlsx_bytes(
        output.getvalue(), target_student_reference="SYNTHETIC_TARGET"
    )

    assert len(preview.rows) == 1
    assert preview.rows[0].raw_score == Decimal("91")
    assert preview.rows[0].source_sheet == "합성 학급"
    assert "SYNTHETIC_OTHER" not in repr(preview)
    assert "SYNTHETIC_TARGET" not in repr(preview)
