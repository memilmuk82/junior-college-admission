from __future__ import annotations

import csv
import hashlib
import io
from collections.abc import Mapping
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Literal

from openpyxl import load_workbook

TextSourceFormat = Literal["csv", "pasted_table"]
SourceFormat = Literal["csv", "pasted_table", "xlsx", "text_pdf"]
ScoreValue = Decimal | Literal["P"] | None
MAX_XLSX_BYTES = 20 * 1024 * 1024
MAX_XLSX_SHEETS = 20
MAX_XLSX_HEADER_SCAN_ROWS = 50
MAX_XLSX_DATA_ROWS_PER_SHEET = 5000

HEADER_ALIASES = {
    "학년도": "academic_year",
    "academic_year": "academic_year",
    "학년": "grade",
    "grade": "grade",
    "학기": "semester",
    "semester": "semester",
    "교과": "subject_group",
    "교과군": "subject_group",
    "subject_group": "subject_group",
    "과목": "subject_name",
    "과목명": "subject_name",
    "subject_name": "subject_name",
    "학점": "credits",
    "이수단위": "credits",
    "credits": "credits",
    "원점수": "raw_score",
    "raw_score": "raw_score",
    "평균": "course_mean",
    "course_mean": "course_mean",
    "표준편차": "standard_deviation",
    "standard_deviation": "standard_deviation",
    "성취도": "achievement_level",
    "achievement_level": "achievement_level",
    "수강자수": "enrollment_count",
    "수강자 수": "enrollment_count",
    "enrollment_count": "enrollment_count",
    "석차등급": "rank_grade",
    "rank_grade": "rank_grade",
}
STUDENT_REFERENCE_HEADERS = {
    "학생식별자",
    "학생 식별자",
    "student_reference",
    "student_ref",
}


@dataclass(frozen=True, slots=True)
class NormalizationIssue:
    code: str
    field: str
    row_number: int
    sheet_name: str | None = None
    source_page: int | None = None


@dataclass(frozen=True, slots=True)
class NormalizedCourseRow:
    academic_year: int | None
    grade: int | None
    semester: int | None
    subject_group: str | None
    subject_name: str | None
    credits: Decimal | None
    raw_score: ScoreValue
    course_mean: Decimal | None
    standard_deviation: Decimal | None
    achievement_level: str | None
    enrollment_count: int | None
    rank_grade: Decimal | None
    source_sheet: str | None = None
    source_row_number: int | None = None
    source_page: int | None = None


@dataclass(frozen=True, slots=True)
class StructuredImportPreview:
    source_hash: str
    source_format: SourceFormat
    rows: tuple[NormalizedCourseRow, ...]
    issues: tuple[NormalizationIssue, ...]
    ignored_headers: tuple[str, ...]


class StructuredInputLimitError(ValueError):
    pass


def _clean_text(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


def _parse_decimal(
    value: str | None,
    *,
    field: str,
    row_number: int,
    issues: list[NormalizationIssue],
    sheet_name: str | None = None,
) -> Decimal | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        issues.append(NormalizationIssue("INVALID_DECIMAL", field, row_number, sheet_name))
        return None


def _parse_integer(
    value: str | None,
    *,
    field: str,
    row_number: int,
    issues: list[NormalizationIssue],
    sheet_name: str | None = None,
) -> int | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    try:
        parsed = Decimal(cleaned)
    except InvalidOperation:
        issues.append(NormalizationIssue("INVALID_INTEGER", field, row_number, sheet_name))
        return None
    if parsed != parsed.to_integral_value():
        issues.append(NormalizationIssue("INVALID_INTEGER", field, row_number, sheet_name))
        return None
    return int(parsed)


def _delimiter_for(source: str, source_format: TextSourceFormat) -> str:
    if source_format == "csv":
        return ","
    header = source.splitlines()[0] if source.splitlines() else ""
    if "\t" in header:
        return "\t"
    try:
        return csv.Sniffer().sniff(source, delimiters=",;|").delimiter
    except csv.Error:
        return ","


def _standardized_row(row: Mapping[object, object], header_map: dict[str, str]) -> dict[str, str]:
    standardized: dict[str, str] = {}
    for original_header, value in row.items():
        field = header_map.get(original_header) if isinstance(original_header, str) else None
        if field is not None and isinstance(value, str):
            standardized[field] = value
    return standardized


def _normalize_course_row(
    row: Mapping[str, str],
    *,
    row_number: int,
    issues: list[NormalizationIssue],
    sheet_name: str | None = None,
) -> NormalizedCourseRow:
    subject_name = _clean_text(row.get("subject_name"))
    if subject_name is None:
        issues.append(
            NormalizationIssue("MISSING_REQUIRED", "subject_name", row_number, sheet_name)
        )

    raw_score_text = _clean_text(row.get("raw_score"))
    if raw_score_text == "P":
        raw_score: ScoreValue = "P"
    else:
        raw_score = _parse_decimal(
            raw_score_text,
            field="raw_score",
            row_number=row_number,
            issues=issues,
            sheet_name=sheet_name,
        )

    return NormalizedCourseRow(
        academic_year=_parse_integer(
            row.get("academic_year"),
            field="academic_year",
            row_number=row_number,
            issues=issues,
            sheet_name=sheet_name,
        ),
        grade=_parse_integer(
            row.get("grade"),
            field="grade",
            row_number=row_number,
            issues=issues,
            sheet_name=sheet_name,
        ),
        semester=_parse_integer(
            row.get("semester"),
            field="semester",
            row_number=row_number,
            issues=issues,
            sheet_name=sheet_name,
        ),
        subject_group=_clean_text(row.get("subject_group")),
        subject_name=subject_name,
        credits=_parse_decimal(
            row.get("credits"),
            field="credits",
            row_number=row_number,
            issues=issues,
            sheet_name=sheet_name,
        ),
        raw_score=raw_score,
        course_mean=_parse_decimal(
            row.get("course_mean"),
            field="course_mean",
            row_number=row_number,
            issues=issues,
            sheet_name=sheet_name,
        ),
        standard_deviation=_parse_decimal(
            row.get("standard_deviation"),
            field="standard_deviation",
            row_number=row_number,
            issues=issues,
            sheet_name=sheet_name,
        ),
        achievement_level=_clean_text(row.get("achievement_level")),
        enrollment_count=_parse_integer(
            row.get("enrollment_count"),
            field="enrollment_count",
            row_number=row_number,
            issues=issues,
            sheet_name=sheet_name,
        ),
        rank_grade=_parse_decimal(
            row.get("rank_grade"),
            field="rank_grade",
            row_number=row_number,
            issues=issues,
            sheet_name=sheet_name,
        ),
        source_sheet=sheet_name,
        source_row_number=row_number,
    )


def parse_structured_text(
    source: str, *, source_format: TextSourceFormat
) -> StructuredImportPreview:
    source_hash = hashlib.sha256(source.encode("utf-8")).hexdigest()
    reader = csv.DictReader(io.StringIO(source), delimiter=_delimiter_for(source, source_format))
    raw_headers = tuple(reader.fieldnames or ())
    cleaned_headers = tuple((header or "").removeprefix("\ufeff").strip() for header in raw_headers)
    header_map = {
        raw_header: HEADER_ALIASES[cleaned_header]
        for raw_header, cleaned_header in zip(raw_headers, cleaned_headers, strict=True)
        if cleaned_header in HEADER_ALIASES
    }
    ignored_headers = tuple(
        header for header in cleaned_headers if header and header not in HEADER_ALIASES
    )
    issues: list[NormalizationIssue] = []
    normalized_rows: list[NormalizedCourseRow] = []

    for row_number, raw_row in enumerate(reader, start=2):
        if not any(_clean_text(value) for value in raw_row.values() if isinstance(value, str)):
            continue
        row = _standardized_row(raw_row, header_map)
        normalized_rows.append(_normalize_course_row(row, row_number=row_number, issues=issues))

    return StructuredImportPreview(
        source_hash=source_hash,
        source_format=source_format,
        rows=tuple(normalized_rows),
        issues=tuple(issues),
        ignored_headers=ignored_headers,
    )


def parse_target_student_text(
    source: str,
    *,
    source_format: TextSourceFormat,
    target_student_reference: str,
) -> StructuredImportPreview:
    target_reference = target_student_reference.strip()
    if not target_reference:
        raise ValueError("대상 학생 식별자가 필요합니다.")

    source_hash = hashlib.sha256(source.encode("utf-8")).hexdigest()
    reader = csv.DictReader(io.StringIO(source), delimiter=_delimiter_for(source, source_format))
    raw_headers = tuple(reader.fieldnames or ())
    cleaned_headers = tuple((header or "").removeprefix("\ufeff").strip() for header in raw_headers)
    standard_header_map = {
        raw_header: HEADER_ALIASES[cleaned_header]
        for raw_header, cleaned_header in zip(raw_headers, cleaned_headers, strict=True)
        if cleaned_header in HEADER_ALIASES
    }
    student_reference_header = next(
        (
            raw_header
            for raw_header, cleaned_header in zip(raw_headers, cleaned_headers, strict=True)
            if cleaned_header in STUDENT_REFERENCE_HEADERS
        ),
        None,
    )
    ignored_headers = tuple(
        header
        for header in cleaned_headers
        if header and header not in HEADER_ALIASES and header not in STUDENT_REFERENCE_HEADERS
    )
    issues: list[NormalizationIssue] = []
    normalized_rows: list[NormalizedCourseRow] = []

    if student_reference_header is None:
        issues.append(NormalizationIssue("TARGET_COLUMN_NOT_FOUND", "student_reference", 0))
    else:
        for row_number, raw_row in enumerate(reader, start=2):
            candidate_reference = raw_row.get(student_reference_header)
            if not isinstance(candidate_reference, str):
                continue
            if candidate_reference.strip() != target_reference:
                continue
            row = _standardized_row(raw_row, standard_header_map)
            normalized_rows.append(_normalize_course_row(row, row_number=row_number, issues=issues))

        if not normalized_rows:
            issues.append(NormalizationIssue("TARGET_NOT_FOUND", "student_reference", 0))

    return StructuredImportPreview(
        source_hash=source_hash,
        source_format=source_format,
        rows=tuple(normalized_rows),
        issues=tuple(issues),
        ignored_headers=ignored_headers,
    )


def _xlsx_cell_text(value: object) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _xlsx_header_fields(
    values: tuple[object, ...], *, require_student_reference: bool
) -> tuple[str | None, ...] | None:
    fields: list[str | None] = []
    for value in values:
        cleaned = _xlsx_cell_text(value)
        if cleaned in STUDENT_REFERENCE_HEADERS:
            fields.append("_student_reference")
        else:
            fields.append(HEADER_ALIASES.get(cleaned) if cleaned else None)
    recognized_fields = tuple(field for field in fields if field is not None)
    if "subject_name" not in recognized_fields or len(recognized_fields) < 2:
        return None
    if require_student_reference and "_student_reference" not in recognized_fields:
        return None
    return tuple(fields)


def _parse_xlsx_bytes(
    source: bytes, *, target_student_reference: str | None
) -> StructuredImportPreview:
    if len(source) > MAX_XLSX_BYTES:
        raise StructuredInputLimitError("XLSX 입력 크기 제한을 초과했습니다.")

    workbook = load_workbook(
        filename=io.BytesIO(source),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        if len(workbook.worksheets) > MAX_XLSX_SHEETS:
            raise StructuredInputLimitError("XLSX 시트 수 제한을 초과했습니다.")

        issues: list[NormalizationIssue] = []
        normalized_rows: list[NormalizedCourseRow] = []
        ignored_headers: list[str] = []

        for worksheet in workbook.worksheets:
            header_row_number: int | None = None
            header_values: tuple[object, ...] = ()
            header_fields: tuple[str | None, ...] | None = None

            for candidate_row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=MAX_XLSX_HEADER_SCAN_ROWS,
                    values_only=True,
                ),
                start=1,
            ):
                candidate_values = tuple(values)
                candidate_fields = _xlsx_header_fields(
                    candidate_values,
                    require_student_reference=target_student_reference is not None,
                )
                if candidate_fields is not None:
                    header_row_number = candidate_row_number
                    header_values = candidate_values
                    header_fields = candidate_fields
                    break

            if header_row_number is None or header_fields is None:
                issues.append(NormalizationIssue("HEADER_NOT_FOUND", "", 0, worksheet.title))
                continue

            for value, field in zip(header_values, header_fields, strict=True):
                header = _xlsx_cell_text(value)
                if (
                    header
                    and field is None
                    and header not in STUDENT_REFERENCE_HEADERS
                    and header not in ignored_headers
                ):
                    ignored_headers.append(header)

            data_row_count = 0
            for row_number, values in enumerate(
                worksheet.iter_rows(min_row=header_row_number + 1, values_only=True),
                start=header_row_number + 1,
            ):
                data_row_count += 1
                if data_row_count > MAX_XLSX_DATA_ROWS_PER_SHEET:
                    raise StructuredInputLimitError(
                        f"XLSX 시트 행 수 제한을 초과했습니다: {worksheet.title}"
                    )
                if not any(_xlsx_cell_text(value) for value in values):
                    continue
                row = {
                    field: cell_text
                    for field, value in zip(header_fields, values, strict=False)
                    if field is not None and (cell_text := _xlsx_cell_text(value)) is not None
                }
                if target_student_reference is not None:
                    candidate_reference = row.pop("_student_reference", None)
                    if candidate_reference != target_student_reference:
                        continue
                normalized_rows.append(
                    _normalize_course_row(
                        row,
                        row_number=row_number,
                        issues=issues,
                        sheet_name=worksheet.title,
                    )
                )

        if target_student_reference is not None and not normalized_rows:
            issues.append(NormalizationIssue("TARGET_NOT_FOUND", "student_reference", 0))

        return StructuredImportPreview(
            source_hash=hashlib.sha256(source).hexdigest(),
            source_format="xlsx",
            rows=tuple(normalized_rows),
            issues=tuple(issues),
            ignored_headers=tuple(ignored_headers),
        )
    finally:
        workbook.close()


def parse_xlsx_bytes(source: bytes) -> StructuredImportPreview:
    return _parse_xlsx_bytes(source, target_student_reference=None)


def parse_target_student_xlsx_bytes(
    source: bytes, *, target_student_reference: str
) -> StructuredImportPreview:
    target_reference = target_student_reference.strip()
    if not target_reference:
        raise ValueError("대상 학생 식별자가 필요합니다.")
    return _parse_xlsx_bytes(source, target_student_reference=target_reference)
