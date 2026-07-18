from __future__ import annotations

import csv
import hashlib
import json
import re
from collections.abc import Iterator, Mapping, Sequence
from dataclasses import dataclass, replace
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from typing import Protocol
from unicodedata import normalize

from openpyxl import load_workbook

MAX_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_SHEETS = 30
MAX_ROWS = 100_000
HEADER_SCAN_ROWS = 50
DEFAULT_COLUMN_ALIAS_PATH = (
    Path(__file__).resolve().parents[2]
    / "data"
    / "seed"
    / "phase14_admission_result_column_aliases.json"
)
ALIAS_SCHEMA = "phase14_admission_result_column_aliases"
ALIAS_VERSION = 1
_CANONICAL_COLUMNS = frozenset(
    {
        "result_academic_year",
        "region",
        "institution_name",
        "campus_name",
        "admission_round_name",
        "program_name",
        "day_night",
        "admission_category",
        "admission_track_name",
        "capacity",
        "applicant_count",
        "admitted_count",
        "competition_rate",
        "best_score",
        "average_score",
        "cutoff_score",
        "score_basis",
        "score_direction",
        "source_reference",
    }
)


class AdmissionResultUploadError(ValueError):
    pass


def load_admission_result_column_aliases(path: Path | None = None) -> dict[str, str]:
    config_path = path or DEFAULT_COLUMN_ALIAS_PATH
    try:
        payload: object = json.loads(config_path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as error:
        raise AdmissionResultUploadError("입시결과 열 동의어 설정을 읽을 수 없습니다.") from error
    if not isinstance(payload, dict):
        raise AdmissionResultUploadError("입시결과 열 동의어 설정의 최상위 구조가 잘못되었습니다.")
    if payload.get("schema") != ALIAS_SCHEMA or payload.get("version") != ALIAS_VERSION:
        raise AdmissionResultUploadError("입시결과 열 동의어 schema 또는 version이 다릅니다.")
    configured = payload.get("aliases")
    if not isinstance(configured, dict) or set(configured) != _CANONICAL_COLUMNS:
        raise AdmissionResultUploadError("입시결과 열 동의어 canonical 필드 구성이 잘못되었습니다.")

    lookup: dict[str, str] = {}
    for canonical, aliases in configured.items():
        if not isinstance(canonical, str) or not isinstance(aliases, list) or not aliases:
            raise AdmissionResultUploadError("입시결과 열 동의어 항목 형식이 잘못되었습니다.")
        for alias in aliases:
            if not isinstance(alias, str) or not alias.strip():
                raise AdmissionResultUploadError("입시결과 열 동의어는 빈 문자열일 수 없습니다.")
            normalized = _normalize_header(alias)
            if not normalized:
                raise AdmissionResultUploadError("정규화 결과가 빈 열 동의어는 사용할 수 없습니다.")
            previous = lookup.get(normalized)
            if previous is not None and previous != canonical:
                raise AdmissionResultUploadError(
                    "정규화된 열 동의어가 둘 이상의 canonical 필드와 충돌합니다."
                )
            lookup[normalized] = canonical
    return lookup


@dataclass(frozen=True)
class CatalogMatch:
    institution_code: str
    campus_code: str
    program_code: str
    admission_round_code: str
    admission_track_code: str
    campus_name: str


class CatalogResolver(Protocol):
    def resolve(
        self,
        *,
        institution_name: str,
        campus_name: str | None,
        program_name: str,
        admission_round_name: str,
        admission_track_name: str,
        target_academic_year: int,
    ) -> CatalogMatch | None: ...


@dataclass(frozen=True)
class ImportIssue:
    code: str
    message: str


@dataclass(frozen=True)
class CanonicalAdmissionResultRow:
    source_sheet: str
    source_row_number: int
    result_academic_year: int
    target_academic_year: int
    region: str | None
    institution_code: str | None
    institution_name: str
    campus_code: str | None
    campus_name: str | None
    admission_round_code: str | None
    admission_round_name: str
    program_code: str | None
    program_name: str
    day_night: str | None
    admission_category: str | None
    admission_track_code: str | None
    admission_track_name: str
    capacity: int | None
    applicant_count: int | None
    admitted_count: int | None
    competition_rate: Decimal | None
    best_score: Decimal | None
    average_score: Decimal | None
    cutoff_score: Decimal | None
    score_basis: str
    score_direction: str
    historical_score_rule_id: str | None
    historical_score_rule_version: str | None
    historical_score_rule_year: int | None
    source_reference: str
    validation_status: str
    issues: tuple[ImportIssue, ...]

    @property
    def business_key(self) -> tuple[object, ...] | None:
        values = (
            self.target_academic_year,
            self.institution_code,
            self.campus_code,
            self.program_code,
            self.admission_round_code,
            self.admission_track_code,
            self.score_basis,
        )
        return values if all(value is not None and value != "" for value in values) else None


@dataclass(frozen=True)
class AdmissionResultUploadPreview:
    source_hash: str
    source_format: str
    result_academic_year: int
    target_academic_year: int
    detected_sheets: tuple[str, ...]
    column_mapping: tuple[tuple[str, str], ...]
    available_source_columns: tuple[str, ...]
    rows: tuple[CanonicalAdmissionResultRow, ...]

    @property
    def total_row_count(self) -> int:
        return len(self.rows)

    @property
    def valid_row_count(self) -> int:
        return sum(row.validation_status == "VALID" for row in self.rows)

    @property
    def review_row_count(self) -> int:
        return sum(row.validation_status == "REVIEW" for row in self.rows)

    @property
    def error_row_count(self) -> int:
        return sum(row.validation_status == "ERROR" for row in self.rows)


def parse_admission_result_upload(
    source: bytes,
    *,
    filename: str,
    result_academic_year: int,
    target_academic_year: int | None = None,
    catalog: CatalogResolver | None = None,
    column_alias_config_path: Path | None = None,
    column_overrides: Mapping[str, str] | None = None,
) -> AdmissionResultUploadPreview:
    if not source or len(source) > MAX_UPLOAD_BYTES:
        raise AdmissionResultUploadError("입시결과 파일은 1 byte 이상 20 MiB 이하여야 합니다.")
    if not 2000 <= result_academic_year <= 2100:
        raise AdmissionResultUploadError("결과 학년도가 유효하지 않습니다.")
    resolved_target_year = (
        result_academic_year + 1 if target_academic_year is None else target_academic_year
    )
    if not 2000 <= resolved_target_year <= 2100:
        raise AdmissionResultUploadError("상담 대상 학년도가 유효하지 않습니다.")

    alias_lookup = load_admission_result_column_aliases(column_alias_config_path)
    normalized_overrides = _validated_column_overrides(column_overrides)
    for canonical, source_column in normalized_overrides.items():
        normalized_source = _normalize_header(source_column)
        previous = alias_lookup.get(normalized_source)
        if previous is not None and previous != canonical:
            del alias_lookup[normalized_source]
        alias_lookup[normalized_source] = canonical
    source_hash = hashlib.sha256(source).hexdigest()
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        source_format = "CSV"
        tables = tuple(_csv_tables(source, alias_lookup, normalized_overrides))
    elif suffix == ".xlsx":
        source_format = "XLSX"
        tables = tuple(_xlsx_tables(source, alias_lookup, normalized_overrides))
    else:
        raise AdmissionResultUploadError("입시결과 파일은 CSV 또는 XLSX만 지원합니다.")
    if not tables:
        raise AdmissionResultUploadError("입시결과 머리글이 있는 시트를 찾지 못했습니다.")

    rows: list[CanonicalAdmissionResultRow] = []
    mappings: dict[str, str] = {}
    for table in tables:
        mappings.update(table.mapping)
        for row_number, values in table.rows:
            parsed = _canonical_row(
                values,
                source_sheet=table.name,
                source_row_number=row_number,
                result_academic_year=result_academic_year,
                target_academic_year=resolved_target_year,
                source_hash=source_hash,
                catalog=catalog,
            )
            if parsed is not None:
                rows.append(parsed)
                if len(rows) > MAX_ROWS:
                    raise AdmissionResultUploadError("입시결과 행 수가 안전 제한을 초과했습니다.")
    if not rows:
        raise AdmissionResultUploadError("입시결과 데이터 행이 비어 있습니다.")

    rows = _mark_duplicates(rows)
    return AdmissionResultUploadPreview(
        source_hash=source_hash,
        source_format=source_format,
        result_academic_year=result_academic_year,
        target_academic_year=resolved_target_year,
        detected_sheets=tuple(table.name for table in tables),
        column_mapping=tuple(sorted(mappings.items())),
        available_source_columns=tuple(
            dict.fromkeys(column for table in tables for column in table.available_columns)
        ),
        rows=tuple(rows),
    )


@dataclass(frozen=True)
class _Table:
    name: str
    mapping: Mapping[str, str]
    available_columns: tuple[str, ...]
    rows: tuple[tuple[int, Mapping[str, object]], ...]


def _csv_tables(
    source: bytes,
    alias_lookup: Mapping[str, str],
    column_overrides: Mapping[str, str],
) -> Iterator[_Table]:
    try:
        text = source.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise AdmissionResultUploadError("CSV는 UTF-8 형식이어야 합니다.") from error
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(StringIO(text), dialect))
    detected = _detect_header(rows, alias_lookup, column_overrides)
    if detected is None:
        return
    header_index, indexes, mapping, available_columns = detected
    yield _table_from_rows("CSV", rows, header_index, indexes, mapping, available_columns)


def _xlsx_tables(
    source: bytes,
    alias_lookup: Mapping[str, str],
    column_overrides: Mapping[str, str],
) -> Iterator[_Table]:
    try:
        workbook = load_workbook(BytesIO(source), read_only=True, data_only=False)
    except Exception as error:
        raise AdmissionResultUploadError("XLSX 구조를 읽을 수 없습니다.") from error
    try:
        if len(workbook.sheetnames) > MAX_SHEETS:
            raise AdmissionResultUploadError("XLSX 시트 수가 안전 제한을 초과했습니다.")
        for worksheet in workbook.worksheets:
            scanned = [
                [cell.value for cell in row]
                for row in worksheet.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS)
            ]
            detected = _detect_header(scanned, alias_lookup, column_overrides)
            if detected is None:
                continue
            header_index, indexes, mapping, available_columns = detected
            raw_rows: list[Sequence[object]] = []
            raw_rows.extend(scanned)
            raw_rows.extend(
                [cell.value for cell in row]
                for row in worksheet.iter_rows(min_row=HEADER_SCAN_ROWS + 1)
            )
            yield _table_from_rows(
                worksheet.title,
                raw_rows,
                header_index,
                indexes,
                mapping,
                available_columns,
            )
    finally:
        workbook.close()


def _detect_header(
    rows: Sequence[Sequence[object]],
    alias_lookup: Mapping[str, str],
    column_overrides: Mapping[str, str],
) -> tuple[int, dict[int, str], dict[str, str], tuple[str, ...]] | None:
    for row_index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        indexes: dict[int, str] = {}
        mapping: dict[str, str] = {}
        for column_index, value in enumerate(row):
            normalized = _normalize_header(value)
            canonical = alias_lookup.get(normalized)
            selected_override = column_overrides.get(canonical or "")
            if selected_override is not None and normalized != _normalize_header(selected_override):
                continue
            if canonical is not None and canonical not in indexes.values():
                indexes[column_index] = canonical
                mapping[canonical] = str(value).strip()
        found = set(indexes.values())
        if {
            "institution_name",
            "admission_round_name",
            "program_name",
            "admission_track_name",
        }.issubset(found) and found.intersection(
            {"average_score", "cutoff_score", "competition_rate"}
        ):
            available_columns = tuple(
                str(value).strip() for value in row if value is not None and str(value).strip()
            )
            return row_index, indexes, mapping, available_columns
    return None


def _table_from_rows(
    name: str,
    rows: Sequence[Sequence[object]],
    header_index: int,
    indexes: Mapping[int, str],
    mapping: Mapping[str, str],
    available_columns: tuple[str, ...],
) -> _Table:
    converted: list[tuple[int, Mapping[str, object]]] = []
    for row_number, raw in enumerate(rows[header_index + 1 :], header_index + 2):
        values = {
            canonical: raw[index] if index < len(raw) else None
            for index, canonical in indexes.items()
        }
        if any(value not in (None, "") for value in values.values()):
            converted.append((row_number, values))
    return _Table(name, mapping, available_columns, tuple(converted))


def _canonical_row(
    values: Mapping[str, object],
    *,
    source_sheet: str,
    source_row_number: int,
    result_academic_year: int,
    target_academic_year: int,
    source_hash: str,
    catalog: CatalogResolver | None,
) -> CanonicalAdmissionResultRow | None:
    issues: list[ImportIssue] = []
    try:
        institution_name = _text(values.get("institution_name"), required=True)
        institution_name = _normalize_institution_name(institution_name or "")
        campus_name = _text(values.get("campus_name"))
        program_name = _text(values.get("program_name"), required=True) or ""
        admission_round_name = _text(values.get("admission_round_name"), required=True) or ""
        category = _text(values.get("admission_category"))
        track = _text(values.get("admission_track_name"), required=True) or ""
        admission_track_name = " / ".join(value for value in (category, track) if value)
        region = _text(values.get("region"))
        day_night = _text(values.get("day_night"))
        score_basis = (_text(values.get("score_basis")) or "RANK_GRADE").upper()
        score_direction = (
            _text(values.get("score_direction"))
            or ("LOWER_IS_BETTER" if score_basis == "RANK_GRADE" else "HIGHER_IS_BETTER")
        ).upper()
        explicit_source_reference = _text(values.get("source_reference"))
        source_result_year = _integer(values.get("result_academic_year"), "모집학년도")
        capacity = _integer(values.get("capacity"), "모집인원")
        applicant_count = _integer(values.get("applicant_count"), "지원자수")
        admitted_count = _integer(values.get("admitted_count"), "합격자수")
        competition_rate = _decimal(values.get("competition_rate"), "경쟁률")
        best_score = _decimal(values.get("best_score"), "최고 성적")
        average_score = _decimal(values.get("average_score"), "평균 성적")
        cutoff_score = _decimal(values.get("cutoff_score"), "최저 성적")
    except _FormulaError:
        issues.append(
            ImportIssue("FORMULA_NOT_ALLOWED", "수식형 셀은 실행하거나 가져오지 않습니다.")
        )
        return _error_row(
            source_sheet,
            source_row_number,
            result_academic_year,
            target_academic_year,
            source_hash,
            tuple(issues),
        )
    except ValueError as error:
        issues.append(ImportIssue("INVALID_VALUE", str(error)))
        return _error_row(
            source_sheet,
            source_row_number,
            result_academic_year,
            target_academic_year,
            source_hash,
            tuple(issues),
        )

    if all(
        value is None
        for value in (
            capacity,
            applicant_count,
            admitted_count,
            competition_rate,
            best_score,
            average_score,
            cutoff_score,
        )
    ):
        issues.append(ImportIssue("RESULT_METRIC_MISSING", "입시결과 지표가 모두 비어 있습니다."))
    if source_result_year is not None and source_result_year != result_academic_year:
        issues.append(
            ImportIssue(
                "RESULT_YEAR_MISMATCH",
                "행의 모집학년도가 선택한 결과 학년도와 다릅니다.",
            )
        )

    if score_basis not in {"RANK_GRADE", "POINT_SCORE"}:
        issues.append(
            ImportIssue("INVALID_SCORE_BASIS", "지원하지 않는 점수 척도는 게시할 수 없습니다.")
        )
    elif score_basis == "RANK_GRADE":
        if score_direction != "LOWER_IS_BETTER":
            issues.append(
                ImportIssue(
                    "SCORE_DIRECTION_MISMATCH",
                    "석차등급은 낮을수록 좋은 방향이어야 합니다.",
                )
            )
        rank_values = (best_score, average_score, cutoff_score)
        if any(
            value is not None and not Decimal("1") <= value <= Decimal("9") for value in rank_values
        ):
            issues.append(
                ImportIssue("SCORE_OUT_OF_RANGE", "석차등급 결과는 1 이상 9 이하여야 합니다.")
            )
    else:
        issues.append(
            ImportIssue(
                "SCORE_BASIS_REVIEW_REQUIRED",
                "배점 척도는 등급 결과와 직접 비교하지 않고 별도 검수가 필요합니다.",
            )
        )
        if score_direction != "HIGHER_IS_BETTER":
            issues.append(
                ImportIssue(
                    "SCORE_DIRECTION_MISMATCH",
                    "배점 척도의 점수 방향을 확인해야 합니다.",
                )
            )

    match = (
        catalog.resolve(
            institution_name=institution_name,
            campus_name=campus_name,
            program_name=program_name,
            admission_round_name=admission_round_name,
            admission_track_name=admission_track_name,
            target_academic_year=target_academic_year,
        )
        if catalog is not None
        else None
    )
    if match is None:
        issues.append(
            ImportIssue(
                "CATALOG_MAPPING_REQUIRED",
                "대학·캠퍼스·학과·모집시기·전형 업무키를 관리자 검수로 연결해야 합니다.",
            )
        )
    error_codes = {
        "FORMULA_NOT_ALLOWED",
        "INVALID_SCORE_BASIS",
        "INVALID_VALUE",
        "RESULT_METRIC_MISSING",
        "RESULT_YEAR_MISMATCH",
        "SCORE_DIRECTION_MISMATCH",
        "SCORE_OUT_OF_RANGE",
    }
    status = (
        "ERROR"
        if any(issue.code in error_codes for issue in issues)
        else ("REVIEW" if issues else "VALID")
    )
    return CanonicalAdmissionResultRow(
        source_sheet=source_sheet,
        source_row_number=source_row_number,
        result_academic_year=result_academic_year,
        target_academic_year=target_academic_year,
        region=region,
        institution_code=match.institution_code if match else None,
        institution_name=institution_name,
        campus_code=match.campus_code if match else None,
        campus_name=match.campus_name if match else campus_name,
        admission_round_code=match.admission_round_code if match else None,
        admission_round_name=admission_round_name,
        program_code=match.program_code if match else None,
        program_name=program_name,
        day_night=day_night,
        admission_category=category,
        admission_track_code=match.admission_track_code if match else None,
        admission_track_name=admission_track_name,
        capacity=capacity,
        applicant_count=applicant_count,
        admitted_count=admitted_count,
        competition_rate=competition_rate,
        best_score=best_score,
        average_score=average_score,
        cutoff_score=cutoff_score,
        score_basis=score_basis,
        score_direction=score_direction,
        historical_score_rule_id=None,
        historical_score_rule_version=None,
        historical_score_rule_year=None,
        source_reference=explicit_source_reference
        or f"sha256:{source_hash}#{source_sheet}:{source_row_number}",
        validation_status=status,
        issues=tuple(issues),
    )


def _error_row(
    source_sheet: str,
    source_row_number: int,
    result_year: int,
    target_year: int,
    source_hash: str,
    issues: tuple[ImportIssue, ...],
) -> CanonicalAdmissionResultRow:
    return CanonicalAdmissionResultRow(
        source_sheet=source_sheet,
        source_row_number=source_row_number,
        result_academic_year=result_year,
        target_academic_year=target_year,
        region=None,
        institution_code=None,
        institution_name="",
        campus_code=None,
        campus_name=None,
        admission_round_code=None,
        admission_round_name="",
        program_code=None,
        program_name="",
        day_night=None,
        admission_category=None,
        admission_track_code=None,
        admission_track_name="",
        capacity=None,
        applicant_count=None,
        admitted_count=None,
        competition_rate=None,
        best_score=None,
        average_score=None,
        cutoff_score=None,
        score_basis="RANK_GRADE",
        score_direction="LOWER_IS_BETTER",
        historical_score_rule_id=None,
        historical_score_rule_version=None,
        historical_score_rule_year=None,
        source_reference=f"sha256:{source_hash}#{source_sheet}:{source_row_number}",
        validation_status="ERROR",
        issues=issues,
    )


def _mark_duplicates(
    rows: list[CanonicalAdmissionResultRow],
) -> list[CanonicalAdmissionResultRow]:
    grouped: dict[tuple[object, ...], list[int]] = {}
    for index, row in enumerate(rows):
        if row.business_key is not None:
            grouped.setdefault(row.business_key, []).append(index)
    duplicate_indexes = {
        index for indexes in grouped.values() if len(indexes) > 1 for index in indexes
    }
    issue = ImportIssue(
        "DUPLICATE_BUSINESS_KEY", "같은 canonical 업무키가 파일 안에서 중복되었습니다."
    )
    return [
        replace(
            row,
            validation_status="REVIEW"
            if row.validation_status == "VALID"
            else row.validation_status,
            issues=row.issues + (issue,),
        )
        if index in duplicate_indexes
        else row
        for index, row in enumerate(rows)
    ]


def _validated_column_overrides(
    overrides: Mapping[str, str] | None,
) -> dict[str, str]:
    if overrides is None:
        return {}
    validated: dict[str, str] = {}
    claimed_sources: dict[str, str] = {}
    for canonical, source_column in overrides.items():
        if canonical not in _CANONICAL_COLUMNS:
            raise AdmissionResultUploadError("지원하지 않는 canonical 열 override입니다.")
        if not isinstance(source_column, str) or not source_column.strip():
            raise AdmissionResultUploadError("열 override의 source 열은 비어 있을 수 없습니다.")
        source_column = source_column.strip()
        normalized_source = _normalize_header(source_column)
        previous = claimed_sources.get(normalized_source)
        if previous is not None and previous != canonical:
            raise AdmissionResultUploadError(
                "하나의 source 열을 여러 canonical 열에 지정할 수 없습니다."
            )
        claimed_sources[normalized_source] = canonical
        validated[canonical] = source_column
    return validated


class _FormulaError(ValueError):
    pass


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^0-9A-Za-z가-힣]", "", normalize("NFKC", str(value))).lower()


def _formula_like(value: str) -> bool:
    stripped = value.lstrip()
    return (
        bool(stripped)
        and stripped[0] in "=+@"
        or (stripped.startswith("-") and not re.fullmatch(r"-\d+(?:\.\d+)?", stripped))
    )


def _text(value: object, *, required: bool = False) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            raise ValueError("필수 업무키 값이 비어 있습니다.")
        return None
    text = normalize("NFKC", str(value)).strip()
    if _formula_like(text):
        raise _FormulaError
    return text


def _integer(value: object, label: str) -> int | None:
    number = _decimal(value, label)
    if number is None:
        return None
    if number != number.to_integral_value() or number < 0:
        raise ValueError(f"{label}은 0 이상의 정수여야 합니다.")
    return int(number)


def _decimal(value: object, label: str) -> Decimal | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, str) and _formula_like(value):
        raise _FormulaError
    try:
        number = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"{label}이 유효한 숫자가 아닙니다.") from error
    if not number.is_finite() or number < 0:
        raise ValueError(f"{label}은 0 이상의 유한값이어야 합니다.")
    return number


def _normalize_institution_name(value: str) -> str:
    # 제공 자료의 '(특)'은 별도 대학이 아니라 같은 대학의 전형 구분 표기다.
    return re.sub(r"\s*\(특\)\s*$", "", value).strip()


__all__ = [
    "AdmissionResultUploadError",
    "AdmissionResultUploadPreview",
    "CanonicalAdmissionResultRow",
    "CatalogMatch",
    "CatalogResolver",
    "ImportIssue",
    "load_admission_result_column_aliases",
    "parse_admission_result_upload",
]
